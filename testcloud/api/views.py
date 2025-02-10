from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence.models import AnalyzeResult, AnalyzeDocumentRequest
from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions, ContentSettings
from django.http import HttpResponse
from django.middleware.csrf import get_token
from rest_framework.authentication import SessionAuthentication
from datetime import datetime, timedelta
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Q
from django.core.files.storage import default_storage
import base64
import requests
from django.core.files.base import ContentFile
from django.shortcuts import render
from rest_framework import viewsets, filters, status, permissions
from rest_framework.views import APIView
from django.contrib.auth import authenticate, login
from django.views import View
from django.http import JsonResponse
from rest_framework.response import Response
from rest_framework import status, generics
from django.views.decorators.csrf import csrf_exempt
from rest_framework.permissions import IsAuthenticated, AllowAny
from .models import Income, Expense, Budget, Receipt, User, CategoryChoices
from .serializers import (
    UserCreateSerializer,
    UserSerializer,
    IncomeSerializer,
    ExpenseSerializer,
    ReceiptSerializer,
    BudgetSerializer,
    UserSerializer,
    UserCreateSerializer
)
from django.shortcuts import get_object_or_404
from django_filters import rest_framework as filters
from django.contrib.auth import get_user_model, authenticate, login, logout
from django.conf import settings
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import BasePermission, AllowAny
from PIL import Image
from django.utils.decorators import method_decorator
import os
import io
from io import BytesIO
import time
import csv
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import json

User = get_user_model()

class UserViewSet(viewsets.ModelViewSet):
    """
    A ViewSet for retrieving and creating users.
    """
    queryset = User.objects.all()
    serializer_class = UserCreateSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()

        # Generate JWT token for the new user
        refresh = RefreshToken.for_user(user)
        return Response({
            "user": UserSerializer(user).data,
            "refresh": str(refresh),
            "access": str(refresh.access_token),
        }, status=status.HTTP_201_CREATED)

class UserScopedViewSet(viewsets.ModelViewSet):
    """
    A base ViewSet to restrict queryset to the authenticated user.
    """
    permission_classes = [AllowAny]

    def get_queryset(self):
        """
        Limit queryset to objects belonging to the authenticated user.
        """
        return self.queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        """
        Automatically associate the authenticated user with the created object.
        """
        serializer.save(user=self.request.user)

class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        # Get email and password from request data
        email = request.data.get('email')
        password = request.data.get('password')

        # Authenticate the user
        user = authenticate(email=email, password=password)

        if user:
            # Generate JWT token
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'user': {
                    'id': user.id,
                    'email': user.email,
                    'full_name': user.full_name,
                }
            }, status=status.HTTP_200_OK)

        return Response({"error": "Invalid credentials"}, status=status.HTTP_401_UNAUTHORIZED)
# Income ViewSet
#class IncomeViewSet(UserScopedViewSet):
class IncomeViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = IncomeSerializer
    queryset = Income.objects.all()
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ['category', 'date', 'source']
    
    def perform_create(self, serializer):

        #serializer.save(user=self.request.user)
        serializer.save()

# Expense ViewSet
#class ExpenseViewSet(UserScopedViewSet):c
class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self):
        return Expense.objects.filter(user=self.request.user)    
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ['category', 'date', 'vendor', 'payment_method']
    
    def perform_create(self, serializer):

        #serializer.save(user=self.request.user)
        serializer.save(user=self.request.user)

# Budget ViewSet
class BudgetViewSet(viewsets.ModelViewSet):
    serializer_class = BudgetSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self):
        return Budget.objects.filter(user=self.request.user)    
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ['id', 'start_date', 'end_date']

    def perform_create(self, serializer):
        """
        Ensure that `current_spending` is initialized to zero upon creation.
        """
        serializer.save(user=self.request.user)

def _format_price(price_dict):
    if price_dict is None:
        return "N/A"
    return "".join([f"{p}" for p in price_dict.values()])

class ReceiptViewSet(viewsets.ModelViewSet):
    serializer_class = ReceiptSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self):
        return Receipt.objects.filter(user=self.request.user)
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ['id','merchant']
    
    def perform_create(self, serializer):

        serializer.save(user=self.request.user)

def generate_filename(filename):
    timestamp = int(time.time())  # Get current timestamp
    extension = filename.split('.')[-1]  # Extract file extension
    return f"{timestamp}.{extension}"  # Return unique filename

class ProcessReceiptView(APIView):
    #authentication_classes = [SessionAuthentication]  # No CSRF required
    permission_classes = [IsAuthenticated]
    def post(self, request, *args, **kwargs):
        
        image_url = request.data.get('image_url')
        uploaded_file = request.FILES.get('image')
        
        if uploaded_file:
            blob_name = generate_filename(uploaded_file.name)
            receiptUrl = upload_image_to_azure(uploaded_file, blob_name)
        elif image_url:
            try:
                # Download the image from the URL
                response = requests.get(image_url, timeout=10)
                if response.status_code != 200:
                    return Response({"error": "Failed to download image from URL."}, status=status.HTTP_400_BAD_REQUEST)

                # Convert to Django file and upload to Blob Storage
                image_content = ContentFile(response.content)
                blob_name = generate_filename("receipt_from_url.jpg")  # Naming for URL images
                receiptUrl = upload_image_to_azure(image_content, blob_name)

            except requests.exceptions.RequestException as e:
                return Response({"error": f"Error fetching image from URL: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)        
        else:
            return Response({"error": "No image file or image_url provided."}, status=status.HTTP_400_BAD_REQUEST)
        #endpoint = os.environ["DOCUMENTINTELLIGENCE_ENDPOINT"]
        endpoint = "https://testcloud-receipt.cognitiveservices.azure.com/"
        #key = os.environ["DOCUMENTINTELLIGENCE_API_KEY"]
        key = "55ZEzXXYdoiuCQykzrZFzTMUxdD4gaw3kqUx8o1U0heIaVoXxu2vJQQJ99BAACi5YpzXJ3w3AAALACOGLe2w"

        document_intelligence_client = DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key))
        if receiptUrl:
            # Process URL
            poller = document_intelligence_client.begin_analyze_document(
                "prebuilt-receipt",
                AnalyzeDocumentRequest(url_source=receiptUrl)
            )
        
        receipts: AnalyzeResult = poller.result()
        
        if receipts.documents:
            for idx, receipt in enumerate(receipts.documents):
                if receipt.fields:
                    merchant_name = receipt.fields.get("MerchantName")
                    total = receipt.fields.get("Total")
                    items = receipt.fields.get("Items")
                    transaction_date_field = receipt.fields.get("TransactionDate")
                    receipt_category = receipt.fields.get("ReceiptType")
                    receipt_items = []
                    if items:
                        for idx, item in enumerate(items.get("valueArray")):
                            item_details = {}
                            item_description = item.get("valueObject").get("Description")
                            if item_description:
                                item_details["description"] = {
                                "value": item_description.get("valueString"),
                                }

                            item_quantity = item.get("valueObject").get("Quantity")
                            if item_quantity:
                                item_details["quantity"] = {
                                "value": item_quantity.get("valueString"),
                                }

                            item_total_price = item.get("valueObject").get("TotalPrice")
                            if item_total_price:
                                item_details["total_price"] = {
                                "value": str(item_total_price.get("valueCurrency").get("amount")),
                                }
                                category = receipt_category.get('valueString')
                                category_choices = {c.value.lower(): c.value for c in CategoryChoices}
                                assigned_category = category_choices.get(category.lower(), CategoryChoices.OTHER)
                            Expense.objects.create(
                                user=request.user,
                                amount=item_total_price.get("valueCurrency").get("amount"),
                                category=assigned_category,
                                date=transaction_date_field.get("valueDate") if transaction_date_field else None,
                                vendor=merchant_name.get('valueString') if merchant_name else "Unknown Merchant",
                                payment_method=None,
                            )
                            receipt_items.append(item_details)
        category = receipt_category.get('valueString').split(".")[0]
        category_choices = {c.value.lower(): c.value for c in CategoryChoices}
        assigned_category = category_choices.get(category.lower(), CategoryChoices.OTHER)
        receipt = Receipt.objects.create(user=request.user,
                                         image_url=receiptUrl if receiptUrl else None,
                                         merchant=merchant_name.get('valueString') if merchant_name else "Unknown Merchant",
                                         total_amount = float(total.get("valueCurrency", {}).get("amount")) if total else 0.00,
                                         parsed_items=receipt_items,
                                         transaction_date=transaction_date_field.get("valueDate") if transaction_date_field else None,
                                         receipt_category=assigned_category,
                                         )
        receipt.assign_to_budget()
        serializer = ReceiptSerializer(receipt)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
def upload_image_to_azure(image_file, blob_name):
    """Uploads an image to Azure Blob Storage and returns the URL."""
    
    #AZURE_STORAGE_ACCOUNT_NAME = os.environ["AZURE_STORAGE_ACCOUNT_NAME"]
    AZURE_STORAGE_ACCOUNT_NAME = "testcloudblob"
    #AZURE_STORAGE_ACCOUNT_KEY = os.environ["AZURE_STORAGE_ACCOUNT_KEY"]
    AZURE_STORAGE_ACCOUNT_KEY = "EMNuCh/mIyM97+CH6MYiXzeXv7Vwkp2VqAXtn+jaqGfuvZf6biFl4j+4v37b3lwv8JJ0Cplq7E21+AStyNNiKg=="
    #AZURE_CONTAINER_NAME = os.environ["AZURE_CONTAINER_NAME"]
    AZURE_CONTAINER_NAME = "testcloud-blob"


    compressed_image = compress_image(image_file)
    # Connect to Azure Blob Storage
    
    blob_service_client = BlobServiceClient(
        f'https://{AZURE_STORAGE_ACCOUNT_NAME}.blob.core.windows.net',
        credential=AZURE_STORAGE_ACCOUNT_KEY
    )

    blob_client = blob_service_client.get_blob_client(container=AZURE_CONTAINER_NAME, blob=blob_name)
    blob_client.upload_blob(compressed_image, overwrite=True)
    blob_client.set_http_headers(content_settings=ContentSettings(content_type="image/png"))
    # Upload Image

    # Generate SAS URL with expiration time (e.g., 1 hour)
    sas_token = generate_blob_sas(
        account_name=AZURE_STORAGE_ACCOUNT_NAME,
        container_name=AZURE_CONTAINER_NAME,
        blob_name=blob_name,
        account_key=AZURE_STORAGE_ACCOUNT_KEY,
        permission=BlobSasPermissions(read=True),
        expiry=datetime.utcnow() + timedelta(days=365 * 100)  # URL expires in 1 hour
    )

    sas_url = f"{blob_client.url}?{sas_token}"

    return sas_url
    
    
def compress_image(image_file):
    """Compress the image to reduce file size before uploading."""
    img = Image.open(image_file)

    # Convert to JPEG (reduces file size)
    img = img.convert("RGB")

    # Save to buffer with compression
    img_io = BytesIO()
    img.save(img_io, format="JPEG", quality=70)  # Adjust quality (70-80 is good)
    img_io.seek(0)

    return img_io
        
class ExportReceiptsXlsxView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        """ Generate an Excel (.xlsx) file with formatted receipts, either all or by budget """
        
        budget_id = kwargs.get("budget_id")  # Check if budget_id is provided in the URL

        if budget_id:
            # Export receipts for a specific budget
            try:
                budget = Budget.objects.get(id=budget_id, user=request.user)

                receipts = Receipt.objects.filter(
                    Q(transaction_date__range=[budget.start_date, budget.end_date]) |
                    Q(transaction_date__isnull=True, uploaded_at__range=[budget.start_date, budget.end_date]),
                    user=budget.user,
                )
                
                if not receipts.exists():
                    return Response({"message": "No receipts found for this budget."}, status=404)

                filename = f"budget_{budget_id}_receipts.xlsx"
            except Budget.DoesNotExist:
                return Response({"error": "Budget not found"}, status=404)
        else:
            # Export all receipts for the user
            receipts = Receipt.objects.filter(user=request.user)
            filename = "receipts.xlsx"

        # Create an Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Receipts"

        # Define column headers
        headers = ["ID", "Merchant", "Total Amount", "Transaction Date", "Receipt Category", "Item Name", "Item Price"]
        
        ws.append(headers)

        # Format headers
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"].alignment = Alignment(horizontal="center", vertical="center")

        # Write receipt data
        for receipt in receipts:
            transaction_date = receipt.transaction_date.strftime('%d/%m/%Y') if receipt.transaction_date else receipt.uploaded_at.strftime('%d/%m/%Y')

            # First row: Receipt details with the first item
            receipt_row = [
                receipt.id,
                receipt.merchant,
                f"{round(receipt.total_amount, 2):.2f}",
                transaction_date,
                receipt.receipt_category
            ]

            if isinstance(receipt.parsed_items, list) and receipt.parsed_items:
                first_item = receipt.parsed_items[0]
                item_name = first_item.get("description", {}).get("value", "Unknown Item")
                item_price = float(first_item.get("total_price", {}).get("value", "0.00"))
                receipt_row.append(item_name)
                receipt_row.append(f"{item_price:.2f}")
            else:
                receipt_row.append("No Items")
                receipt_row.append("-")

            ws.append(receipt_row)

            # Additional rows for remaining items (without repeating receipt details)
            if isinstance(receipt.parsed_items, list) and len(receipt.parsed_items) > 1:
                for item in receipt.parsed_items[1:]:
                    item_name = item.get("description", {}).get("value", "Unknown Item")
                    item_price = float(item.get("total_price", {}).get("value", "0.00"))
                    ws.append(["", "", "", "", "", item_name, f"{item_price:.2f}"])

        # Auto-adjust column widths
        for col_num, col_cells in enumerate(ws.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_num)
            for cell in col_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Enable wrap text for the Parsed Items column
        # Create HTTP response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response
 
@method_decorator(csrf_exempt, name='dispatch')   
class BudgetReportView(APIView):  
    permission_classes = [IsAuthenticated]

    def get(self, request, budget_id):
        budget = get_object_or_404(Budget, id=budget_id, user=request.user)
        receipts = Receipt.objects.filter(budget=budget)
        
        # Calculate spending summary
        total_spent = sum(receipt.total_amount or 0 for receipt in receipts)
        category_spending = {}
        total_items = 0
        category_items = {}
        
        for receipt in receipts:
            category = receipt.receipt_category
            category_spending[category] = category_spending.get(category, 0) + (receipt.total_amount or 0)
            
            item_count = len(receipt.parsed_items) if receipt.parsed_items else 1
            total_items += item_count
            category_items[category] = category_items.get(category, 0) + item_count
        
        response_data = {
            "budget": BudgetSerializer(budget).data,
            "total_spent": total_spent,
            "category_spending": category_spending,
            "total_items": total_items,
            "category_items": category_items,
            "receipts": ReceiptSerializer(receipts, many=True).data,
        }
        return Response(response_data, status=status.HTTP_200_OK)


class EmailPasswordLoginView(APIView):
    """ Login using email and password, and store session """
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')
        password = request.data.get('password')

        if not email or not password:
            return Response({"error": "Email and password are required."}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(request, email=email, password=password)

        if user:
            return Response({
                "message": "Login successful!",
                "Authorization": f"Basic {base64.b64encode(f'{email}:{password}'.encode()).decode()}"
            }, status=status.HTTP_200_OK)

        return Response({"error": "Invalid email or password"}, status=status.HTTP_401_UNAUTHORIZED)

class LogoutView(APIView):
    """ Logout the user and clear session """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        logout(request)
        return Response({"message": "Logged out successfully!"}, status=status.HTTP_200_OK)
    
class CSRFTokenView(APIView):
    """ API endpoint to provide the CSRF token """
    permission_classes = [AllowAny]  # Allow any user to access this endpoint

    def get(self, request):
        csrf_token = get_token(request)  # Retrieve the CSRF token
        return JsonResponse({"csrfToken": csrf_token})