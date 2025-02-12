from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework import status
from .models import Expense, Receipt, Budget
from datetime import date
from .models import CategoryChoices
from io import BytesIO
from openpyxl import load_workbook

User = get_user_model()

class UserTests(TestCase):
    def setUp(self):
        """Create a test user."""
        self.user = User.objects.create_user(
            email="testuser@example.com",
            password="testpassword",
            full_name="Test User",
            date_of_birth="2000-01-01"
        )
        self.client = APIClient()

    def test_user_creation(self):
        """Ensure user is created successfully."""
        self.assertEqual(self.user.email, "testuser@example.com")

    def test_user_login(self):
        """Ensure user can log in and receive JWT token."""
        response = self.client.post('/api/login/', {
            'email': 'testuser@example.com',
            'password': 'testpassword'
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('access', response.data)

class ExpenseTests(TestCase):
    def setUp(self):
        """Create a test user and an expense."""
        self.user = User.objects.create_user(email="user@example.com", password="password123")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)  # Authenticate user
        
        self.expense = Expense.objects.create(
            user=self.user,
            amount=50.00,
            category="Groceries",
            date=date.today(),
            vendor="Walmart",
            payment_method="Credit Card"
        )

    def test_create_expense(self):
        """Ensure an expense can be created via API."""
        response = self.client.post('/api/expenses/', {
            'amount': 20.00,
            'category': CategoryChoices.TRANSPORTATION.value,
            'date': str(date.today()),
            'vendor': "Uber",
            'payment_method': "Cash"
        })
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_list_expenses(self):
        """Ensure expenses are listed for the authenticated user."""
        response = self.client.get('/api/expenses/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)

'''
class BudgetTests(TestCase):
    def setUp(self):
        """Create a test user and a budget."""
        self.user = User.objects.create_user(email="budgetuser@example.com", password="securepassword")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

        self.budget = Budget.objects.create(
            user=self.user,
            name="Monthly Budget",
            category="Groceries",
            limit_amount=500.00,
            start_date="2024-02-01",
            end_date="2024-02-28"
        )
        self.budget.filter_categories.add(CategoryChoices.MEAL.value, CategoryChoices.ENTERTAINMENT.value)

    def test_create_budget(self):
        """Ensure a budget can be created via API."""
        response = self.client.post('/api/budgets/', {
            'name': "Test Budget",
            'category': CategoryChoices.MEAL.value,
            'limit_amount': 100.00,
            'start_date': "2024-02-01",
            'end_date': "2024-02-28"
        })
        print("Response Data (Budget):", response.data)  # Debugging
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_list_budgets(self):
        """Ensure only the authenticated user's budgets are listed."""
        response = self.client.get('/api/budgets/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        
    
    def test_budget_filtering_receipts(self):
        """Ensure only receipts with the correct category are counted in the budget."""
        # Create an allowed receipt
        receipt1 = Receipt.objects.create(
            user=self.user,
            total_amount=20.00,
            transaction_date="2024-02-10",
            receipt_category=CategoryChoices.MEAL.value
        )
        receipt1.assign_to_budget()

        # Create a non-allowed receipt
        receipt2 = Receipt.objects.create(
            user=self.user,
            total_amount=30.00,
            transaction_date="2024-02-15",
            receipt_category=CategoryChoices.HEALTHCARE.value  # Not in filter categories
        )
        receipt2.assign_to_budget()

        # Refresh budget
        self.budget.refresh_from_db()
        
        # Only receipt1 should be counted
        self.assertEqual(self.budget.current_spending, 20.00)
'''

class ReceiptTests(TestCase):
    def setUp(self):
        """Create a test user and a receipt."""
        self.user = User.objects.create_user(email="receiptuser@example.com", password="securepass")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

        self.receipt = Receipt.objects.create(
            user=self.user,
            image_url="http://example.com/receipt.jpg",
            merchant="McDonald's",
            total_amount=15.99,
            transaction_date="2024-02-01",
            receipt_category="Food",
        )

    def test_create_receipt(self):
        """Ensure a receipt can be created via API."""
        response = self.client.post('/api/receipts/', {
            'image_url': "http://example.com/new_receipt.jpg",
            'merchant': "Starbucks",
            'total_amount': 5.99,
            'transaction_date': "2024-02-02",
            'receipt_category': CategoryChoices.MEAL.value
        })
        print("Response Data (Receipt):", response.data)  # Debugging
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_list_receipts(self):
        """Ensure only the authenticated user's receipts are listed."""
        response = self.client.get('/api/receipts/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)


class ProcessReceiptTests(TestCase):
    def setUp(self):
        """Create a test user for testing receipt processing."""
        self.user = User.objects.create_user(email="processuser@example.com", password="processingpass")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_process_receipt(self):
        """Ensure the receipt processing API works (mocked response)."""
        response = self.client.post('/api/process-receipt/', {
            'image_url': "https://testcloudblob.blob.core.windows.net/testcloud-blob/1739105642.jpg?se=2125-01-16T12%3A54%3A04Z&sp=r&sv=2025-01-05&sr=b&sig=RgVAeIcPvyryJsN3SpPYhsLkjvxdTu8dz71yYo2yK7Y%3D"
        })
        self.assertIn(response.status_code, [status.HTTP_200_OK, status.HTTP_201_CREATED])
        receipt = Receipt.objects.latest('id')
        print("Processed Receipt Category:", receipt.receipt_category)  # Debugging
        
        # ✅ Ensure the category is correctly mapped
        self.assertIn(receipt.receipt_category, CategoryChoices.values)
        
class ExportReceiptsXlsxViewTest(TestCase):
    def setUp(self):
        """Set up a test user, budget, and receipts."""
        self.user = User.objects.create_user(email="xlsxuser@example.com", password="testpass")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

        # ✅ Create a budget
        self.budget = Budget.objects.create(
            user=self.user,
            name="Food Budget",
            category=CategoryChoices.MEAL.value,
            limit_amount=500.00,
            start_date="2024-02-01",
            end_date="2024-02-28"
        )

        # ✅ Create receipts assigned to this budget
        self.receipt1 = Receipt.objects.create(
            user=self.user,
            total_amount=20.00,
            transaction_date="2024-02-10",
            receipt_category=CategoryChoices.MEAL.value
        )
        self.receipt1.budget.add(self.budget)

        self.receipt2 = Receipt.objects.create(
            user=self.user,
            total_amount=30.00,
            transaction_date="2024-02-15",
            receipt_category=CategoryChoices.MEAL.value
        )
        self.receipt2.budget.add(self.budget)

    def test_export_receipts_xlsx(self):
        """Ensure the receipts export returns a valid Excel file."""
        url = f'/api/export/budget/{self.budget.id}/'  # ✅ Adjust to match your URL pattern
        response = self.client.get(url)

        # ✅ Check the response status
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # ✅ Verify the Excel file structure
        excel_data = BytesIO(response.content)
        workbook = load_workbook(excel_data)
        sheet = workbook.active

        # ✅ Check headers
        expected_headers = ["ID", "Merchant", "Total Amount", "Transaction Date", "Receipt Category", "Item Name", "Item Price"]
        sheet_headers = [cell.value for cell in sheet[1]]
        self.assertEqual(sheet_headers, expected_headers)

        # ✅ Check that receipts are included
        self.assertEqual(sheet.cell(row=2, column=3).value, "20.00")  # First receipt amount
        self.assertEqual(sheet.cell(row=3, column=3).value, "30.00")  # Second receipt amount

        print("Exported Receipts:", sheet_headers)  # Debugging
